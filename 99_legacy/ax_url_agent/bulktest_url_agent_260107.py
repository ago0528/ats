"""
에이전트 URL 테스터 - Streamlit GUI
실행: streamlit run bulktest_url_agent_260107.py
"""

import streamlit as st
import asyncio
import aiohttp
import json
import re
import jwt
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from typing import Optional
import time

# ============================================================
# 페이지 설정
# ============================================================
st.set_page_config(
    page_title="에이전트 URL 테스터",
    page_icon="🧪",
    layout="wide"
)

# ============================================================
# 세션 상태 초기화
# ============================================================
if 'results' not in st.session_state:
    st.session_state.results = []
if 'is_running' not in st.session_state:
    st.session_state.is_running = False
if 'cancel_requested' not in st.session_state:
    st.session_state.cancel_requested = False
if 'progress' not in st.session_state:
    st.session_state.progress = 0
if 'total' not in st.session_state:
    st.session_state.total = 0


# ============================================================
# 토큰 검증 함수
# ============================================================
def validate_token(token: str, token_name: str) -> tuple[bool, str]:
    """JWT 토큰 유효성 검증"""
    if not token or not token.strip():
        return False, "토큰이 비어있습니다"
    
    try:
        # Bearer 접두어 제거
        token = token.strip()
        if token.startswith("Bearer "):
            token = token[7:]
        
        payload = jwt.decode(token, options={"verify_signature": False})
        exp = payload.get('exp')
        
        if exp:
            exp_dt = datetime.fromtimestamp(exp)
            now = datetime.now()
            
            if exp_dt < now:
                return False, f"❌ 만료됨 ({exp_dt.strftime('%Y-%m-%d %H:%M:%S')})"
            
            remaining = (exp_dt - now).total_seconds() / 60
            return True, f"✅ 유효 (남은 시간: {remaining:.0f}분)"
        
        return True, "✅ 유효 (만료 시간 없음)"
    
    except Exception as e:
        return False, f"⚠️ 검증 불가: {str(e)}"


# ============================================================
# SSE 파싱 및 테스트 클래스
# ============================================================
class AgentTester:
    def __init__(self, bearer_token: str, cms_token: str, mrs_session: str, max_parallel: int = 1):
        # self.base_url = "https://api-llm.ats.kr-st2-midasin.com"  # QA
        self.base_url = "https://api-llm.ats.kr-pr-midasin.com"  # PR
        self.bearer_token = bearer_token.strip()
        self.cms_token = cms_token.strip()
        self.mrs_session = mrs_session.strip()
        self.max_parallel = max_parallel
        self.semaphore = asyncio.Semaphore(max_parallel)
        self.cancel_flag = False
    
    def get_headers(self, for_sse: bool = False) -> dict:
        headers = {
            "authorization": f"Bearer {self.bearer_token}",
            "cms-access-token": self.cms_token,
            "mrs-session": self.mrs_session,
            # "origin": "https://qa-jobda02-cms-recruiter-co-kr.midasweb.net",  # QA
            # "referer": "https://qa-jobda02-cms-recruiter-co-kr.midasweb.net/",  # QA
            "origin": "https://pr-jobda02-cms.recruiter.co.kr",  # PR
            "referer": "https://pr-jobda02-cms.recruiter.co.kr/",  # PR
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        if for_sse:
            headers["accept"] = "text/event-stream"
        else:
            headers["accept"] = "application/json, text/plain, */*"
            headers["content-type"] = "application/json"
        
        return headers
    
    async def send_query(self, session: aiohttp.ClientSession, message: str) -> tuple[Optional[str], str]:
        """질의 전송 후 (conversationId, 에러메시지) 반환"""
        url = f"{self.base_url}/api/v2/ai/orchestrator/query"
        payload = {
            "conversationId": None,
            "userMessage": message
        }
        
        try:
            async with session.post(url, headers=self.get_headers(), json=payload, timeout=30) as response:
                if response.status == 200:
                    data = await response.json()
                    return data.get("conversationId"), ""
                else:
                    error_text = await response.text()
                    return None, f"HTTP {response.status}: {error_text[:100]}"
        except aiohttp.ClientConnectorError as e:
            return None, f"연결 실패: 네트워크/방화벽 ({str(e)[:50]})"
        except asyncio.TimeoutError:
            return None, "타임아웃: 30초 초과"
        except Exception as e:
            return None, f"{type(e).__name__}: {str(e)[:50]}"
    
    async def subscribe_sse(self, session: aiohttp.ClientSession, conversation_id: str) -> dict:
        """SSE 구독하여 결과 수신"""
        url = f"{self.base_url}/api/v1/ai/orchestrator/chat-room/sse/subscribe"
        params = {"conversationId": conversation_id}
        
        connect_time = None
        chat_time = None
        button_url = None
        error_msg = None
        
        try:
            timeout = aiohttp.ClientTimeout(total=60)
            async with session.get(url, headers=self.get_headers(for_sse=True), 
                                   params=params, timeout=timeout) as response:
                
                buffer = ""
                current_event = None
                last_heartbeat = datetime.now()
                
                async for chunk in response.content.iter_any():
                    # 취소 체크
                    if self.cancel_flag:
                        error_msg = "사용자 취소"
                        break
                    
                    # 하트비트 타임아웃 체크 (30초)
                    if (datetime.now() - last_heartbeat).total_seconds() > 30:
                        error_msg = "응답 타임아웃"
                        break
                    
                    buffer += chunk.decode('utf-8', errors='ignore')
                    
                    while "\n" in buffer:
                        line, buffer = buffer.split("\n", 1)
                        line = line.strip()
                        
                        if line.startswith("event:"):
                            current_event = line.replace("event:", "").strip()
                            
                            if current_event == "CONNECT":
                                connect_time = datetime.now()
                            elif current_event == "HEARTBEAT":
                                last_heartbeat = datetime.now()
                        
                        elif line.startswith("data:"):
                            data_str = line.replace("data:", "", 1).strip()
                            
                            if current_event == "CHAT" and data_str.startswith("{"):
                                try:
                                    data = json.loads(data_str)
                                    if data.get("messageType") == "ASSISTANT":
                                        chat_time = datetime.now()
                                        
                                        # buttonUrl 추출
                                        assistant = data.get("assistant", {})
                                        for ui in assistant.get("dataUIList", []):
                                            ui_value = ui.get("uiValue", {})
                                            if "buttonUrl" in ui_value:
                                                button_url = ui_value["buttonUrl"]
                                                break
                                        
                                        # 결과 반환
                                        response_time = "-"
                                        if connect_time and chat_time:
                                            delta = (chat_time - connect_time).total_seconds()
                                            response_time = f"{delta:.2f}"
                                        
                                        return {
                                            "테스트일시": chat_time.strftime("%Y-%m-%d %H:%M"),
                                            "응답시간(초)": response_time,
                                            "실제URL": button_url or "-",
                                            "실패사유": "" if button_url else "URL 미반환"
                                        }
                                
                                except json.JSONDecodeError:
                                    pass
        
        except asyncio.TimeoutError:
            error_msg = "연결 타임아웃"
        except Exception as e:
            error_msg = str(e)[:50]
        
        return {
            "테스트일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "응답시간(초)": "-",
            "실제URL": "-",
            "실패사유": error_msg or "알 수 없는 오류"
        }
    
    async def run_single_test(self, session: aiohttp.ClientSession, 
                               query_id: str, match: str, query: str, expected_url: str) -> dict:
        """단일 테스트 실행"""
        async with self.semaphore:
            if self.cancel_flag:
                return {
                    "ID": query_id,
                    "매치": match,
                    "질의": query,
                    "기대URL": expected_url,
                    "테스트일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "응답시간(초)": "-",
                    "실제URL": "-",
                    "성공여부": "SKIP",
                    "실패사유": "사용자 취소"
                }
            
            # 1. Query 전송
            conversation_id, query_error = await self.send_query(session, query)
            
            if not conversation_id:
                return {
                    "ID": query_id,
                    "매치": match,
                    "질의": query,
                    "기대URL": expected_url,
                    "테스트일시": datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "응답시간(초)": "-",
                    "실제URL": "-",
                    "성공여부": "FAIL",
                    "실패사유": query_error or "conversationId 획득 실패"
                }
            
            # 2. SSE 구독
            sse_result = await self.subscribe_sse(session, conversation_id)
            
            # 3. 성공 여부 판정
            success, fail_reason = self.evaluate_result(expected_url, sse_result["실제URL"])
            if sse_result["실패사유"]:
                fail_reason = sse_result["실패사유"]
            
            return {
                "ID": query_id,
                "매치": match,
                "질의": query,
                "기대URL": expected_url,
                "테스트일시": sse_result["테스트일시"],
                "응답시간(초)": sse_result["응답시간(초)"],
                "실제URL": sse_result["실제URL"],
                "성공여부": success,
                "실패사유": fail_reason
            }
    
    def evaluate_result(self, expected_url: str, actual_url: str) -> tuple[str, str]:
        """기대 URL과 실제 URL 비교"""
        if not actual_url or actual_url == "-":
            return "FAIL", "URL 미반환"
        
        if not expected_url or expected_url == "-":
            return "PASS", ""  # 기대 URL이 없으면 성공으로 처리
        
        # 패턴 매칭 ({planId}는 와일드카드)
        pattern = expected_url.replace("{planId}", r"(\{planId\}|\d+)")
        pattern = pattern.replace("?", r"\?")
        pattern = pattern.replace("&", r"&")
        
        if re.fullmatch(pattern, actual_url):
            return "PASS", ""
        else:
            return "FAIL", "URL 불일치"
    
    def cancel(self):
        self.cancel_flag = True


# ============================================================
# 질의 파싱 함수
# ============================================================
def parse_queries(text: str) -> list:
    """
    입력 텍스트에서 질의 파싱
    지원 형식:
    1. ID\t매치\t질의\t기대URL (엑셀 원본 형식)
    2. ID\t질의\t기대URL (3컬럼)
    3. ID\t질의 (2컬럼)
    4. 질의만 (자동 ID 부여)
    """
    queries = []
    lines = text.strip().split("\n")
    
    for i, line in enumerate(lines, 1):
        line = line.strip()
        if not line:
            continue
        
        # 헤더 행 스킵
        if line.startswith("ID\t") or line.startswith("ID	"):
            continue
        
        parts = line.split("\t")
        
        if len(parts) >= 4:
            # ID, 매치, 질의, 기대URL (엑셀 원본 형식)
            queries.append({
                "id": parts[0].strip(),
                "match": parts[1].strip(),
                "query": parts[2].strip(),
                "expected_url": parts[3].strip()
            })
        elif len(parts) == 3:
            # ID, 질의, 기대URL (매치 없음)
            queries.append({
                "id": parts[0].strip(),
                "match": "",
                "query": parts[1].strip(),
                "expected_url": parts[2].strip()
            })
        elif len(parts) == 2:
            # ID, 질의
            queries.append({
                "id": parts[0].strip(),
                "match": "",
                "query": parts[1].strip(),
                "expected_url": ""
            })
        else:
            # 질의만
            queries.append({
                "id": f"Q-{i:03d}",
                "match": "",
                "query": line,
                "expected_url": ""
            })
    
    return queries


# ============================================================
# 엑셀 처리 함수
# ============================================================
def read_excel_queries(uploaded_file) -> list:
    """엑셀에서 질의 읽기"""
    queries = []
    
    try:
        wb = load_workbook(uploaded_file)
        ws = wb.active
        
        for row in ws.iter_rows(min_row=2):  # 헤더 제외
            query_id = row[1].value if len(row) > 1 else None  # B열: ID
            query = row[3].value if len(row) > 3 else None      # D열: 질의
            expected_url = row[4].value if len(row) > 4 else "" # E열: 기대 URL
            
            if query_id and query:
                queries.append({
                    "id": str(query_id),
                    "query": str(query),
                    "expected_url": str(expected_url) if expected_url else ""
                })
    except Exception as e:
        st.error(f"엑셀 읽기 오류: {e}")
    
    return queries


def create_result_excel(results: list) -> BytesIO:
    """결과를 엑셀 파일로 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "테스트 결과"
    
    # 헤더
    headers = ["테스트일시", "ID", "매치", "질의", "기대 URL", "실제 URL", "성공 여부", "응답 시간(초)", "실패 사유"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 스타일
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # 데이터
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result.get("테스트일시", ""))
        ws.cell(row=row_idx, column=2, value=result.get("ID", ""))
        ws.cell(row=row_idx, column=3, value=result.get("매치", ""))
        ws.cell(row=row_idx, column=4, value=result.get("질의", ""))
        ws.cell(row=row_idx, column=5, value=result.get("기대URL", ""))
        ws.cell(row=row_idx, column=6, value=result.get("실제URL", ""))
        ws.cell(row=row_idx, column=7, value=result.get("성공여부", ""))
        ws.cell(row=row_idx, column=8, value=result.get("응답시간(초)", ""))
        ws.cell(row=row_idx, column=9, value=result.get("실패사유", ""))
        
        # 성공/실패 색상
        if result.get("성공여부") == "PASS":
            ws.cell(row=row_idx, column=7).fill = pass_fill
        elif result.get("성공여부") == "FAIL":
            ws.cell(row=row_idx, column=7).fill = fail_fill
    
    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    # BytesIO로 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# 비동기 실행 래퍼
# ============================================================
async def run_tests_async(tester: AgentTester, queries: list, progress_callback):
    """비동기 테스트 실행 - 병렬 처리"""
    completed = 0
    
    connector = aiohttp.TCPConnector(limit=tester.max_parallel)
    async with aiohttp.ClientSession(connector=connector) as session:
        
        async def run_with_callback(q):
            nonlocal completed
            result = await tester.run_single_test(
                session, q["id"], q.get("match", ""), q["query"], q["expected_url"]
            )
            completed += 1
            progress_callback(completed, len(queries), result)
            return result
        
        # 병렬 실행 (semaphore가 max_parallel 개수 제한)
        tasks = [run_with_callback(q) for q in queries]
        results = await asyncio.gather(*tasks)
    
    return list(results)


def run_tests_sync(tester: AgentTester, queries: list, progress_placeholder, result_placeholder):
    """동기 래퍼로 비동기 테스트 실행"""
    results = []
    
    def progress_callback(current, total, result):
        progress_placeholder.progress(current / total, f"진행 중... {current}/{total}")
        results.append(result)
        
        # 결과 테이블 업데이트
        df = pd.DataFrame(results)
        display_cols = ["ID", "매치", "질의", "기대URL", "실제URL", "성공여부", "응답시간(초)", "실패사유"]
        display_cols = [c for c in display_cols if c in df.columns]
        result_placeholder.dataframe(df[display_cols], use_container_width=True)
    
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    
    try:
        final_results = loop.run_until_complete(
            run_tests_async(tester, queries, progress_callback)
        )
    finally:
        loop.close()
    
    return final_results


# ============================================================
# Streamlit UI
# ============================================================
def main():
    st.title("🧪 에이전트 URL 테스터")
    
    # ----------------------------------------------------------
    # 사이드바: 설정
    # ----------------------------------------------------------
    with st.sidebar:
        st.header("⚙️ 설정")
        
        bearer_token = st.text_area(
            "Bearer Token",
            height=80,
            placeholder="eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9...",
            key="bearer_token"
        )
        
        cms_token = st.text_area(
            "CMS Access Token",
            height=80,
            placeholder="eyJhbGciOiJSUzI1NiJ9...",
            key="cms_token"
        )
        
        mrs_session = st.text_input(
            "MRS Session",
            placeholder="NTM0MWU2NzQtMWQ0OS00Zjc0...",
            key="mrs_session"
        )
        
        st.divider()
        
        # 토큰 검증 버튼
        if st.button("🔍 토큰 유효성 검사", use_container_width=True):
            if bearer_token:
                valid, msg = validate_token(bearer_token, "Bearer")
                st.write(f"**Bearer Token:** {msg}")
            if cms_token:
                valid, msg = validate_token(cms_token, "CMS")
                st.write(f"**CMS Token:** {msg}")
        
        st.divider()
        
        parallel_count = st.slider("병렬 처리 수", 1, 3, 1)
        
        st.divider()
        
        st.caption("💡 토큰은 크롬 개발자 도구 > Network에서 복사하세요")
    
    # ----------------------------------------------------------
    # 메인: 질의 입력
    # ----------------------------------------------------------
    tab1, tab2 = st.tabs(["📝 직접 입력", "📁 엑셀 업로드"])
    
    with tab1:
        st.markdown("""
        **입력 형식** (탭으로 구분):
        - `ID` `매치` `질의` `기대URL`
        - `ID` `질의` `기대URL`
        - `ID` `질의`
        - `질의` (ID 자동 부여)
        """)
        
        query_text = st.text_area(
            "질의 목록",
            height=200,
            placeholder="CM-01\t기존 채용 불러오기\t이전에 했던 채용을 복사해서 새로 만들고 싶어\t/agent/flow/create?copy={planId}\nCM-02\t기존 채용 불러오기\t기존 채용 불러와서 복제할래\t/agent/flow/create?copy={planId}",
            key="query_text"
        )
    
    with tab2:
        uploaded_file = st.file_uploader(
            "엑셀 파일 업로드 (.xlsx)",
            type=['xlsx'],
            key="excel_upload"
        )
        
        if uploaded_file:
            queries_from_excel = read_excel_queries(uploaded_file)
            st.success(f"✅ {len(queries_from_excel)}개 질의 로드됨")
            
            if queries_from_excel:
                preview_df = pd.DataFrame(queries_from_excel[:5])
                st.dataframe(preview_df, use_container_width=True)
    
    st.divider()
    
    # ----------------------------------------------------------
    # 실행 버튼
    # ----------------------------------------------------------
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        start_button = st.button("▶️ 테스트 시작", type="primary", use_container_width=True)
    
    with col2:
        stop_button = st.button("⏹️ 중지", use_container_width=True)
    
    with col3:
        clear_button = st.button("🗑️ 결과 초기화", use_container_width=True)
    
    if clear_button:
        st.session_state.results = []
        st.rerun()
    
    # 진행률 & 결과 영역
    progress_placeholder = st.empty()
    result_placeholder = st.empty()
    
    # ----------------------------------------------------------
    # 테스트 실행
    # ----------------------------------------------------------
    if start_button:
        # 입력 검증
        if not bearer_token or not cms_token or not mrs_session:
            st.error("❌ 모든 토큰을 입력해주세요")
            return
        
        # 토큰 유효성 체크
        valid, msg = validate_token(bearer_token, "Bearer")
        if not valid:
            st.error(f"❌ Bearer Token 오류: {msg}")
            return
        
        # 질의 파싱
        queries = []
        
        if uploaded_file:
            queries = read_excel_queries(uploaded_file)
        elif query_text:
            queries = parse_queries(query_text)
        
        if not queries:
            st.error("❌ 질의를 입력해주세요")
            return
        
        st.info(f"📋 총 {len(queries)}개 질의 테스트 시작...")
        
        # 테스터 생성 및 실행
        tester = AgentTester(bearer_token, cms_token, mrs_session, parallel_count)
        
        st.session_state.is_running = True
        
        try:
            results = run_tests_sync(tester, queries, progress_placeholder, result_placeholder)
            st.session_state.results = results
            
            # 완료 통계
            pass_count = sum(1 for r in results if r.get("성공여부") == "PASS")
            fail_count = sum(1 for r in results if r.get("성공여부") == "FAIL")
            skip_count = sum(1 for r in results if r.get("성공여부") == "SKIP")
            
            st.success(f"✅ 테스트 완료! PASS: {pass_count} / FAIL: {fail_count} / SKIP: {skip_count}")
        
        except Exception as e:
            st.error(f"❌ 오류 발생: {e}")
            # 중간 결과 저장
            if st.session_state.results:
                st.warning("⚠️ 중간 결과가 저장되었습니다")
        
        finally:
            st.session_state.is_running = False
    
    if stop_button:
        st.session_state.cancel_requested = True
        st.warning("⏹️ 중지 요청됨. 현재 진행 중인 테스트 완료 후 중지됩니다.")
    
    # ----------------------------------------------------------
    # 결과 표시 및 다운로드
    # ----------------------------------------------------------
    if st.session_state.results:
        st.divider()
        st.subheader("📊 테스트 결과")
        
        df = pd.DataFrame(st.session_state.results)
        
        # 통계
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("전체", len(df))
        with col2:
            pass_count = len(df[df["성공여부"] == "PASS"])
            st.metric("PASS", pass_count, delta=None)
        with col3:
            fail_count = len(df[df["성공여부"] == "FAIL"])
            st.metric("FAIL", fail_count, delta=None)
        with col4:
            if len(df) > 0:
                try:
                    avg_time = df[df["응답시간(초)"] != "-"]["응답시간(초)"].astype(float).mean()
                    st.metric("평균 응답시간", f"{avg_time:.2f}초")
                except:
                    st.metric("평균 응답시간", "-")
        
        # 결과 테이블
        st.dataframe(df, use_container_width=True)
        
        # 다운로드 버튼
        col1, col2 = st.columns(2)
        
        with col1:
            excel_data = create_result_excel(st.session_state.results)
            st.download_button(
                label="📥 엑셀 다운로드",
                data=excel_data,
                file_name=f"test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            # 클립보드 복사용 텍스트 (탭 구분)
            clipboard_text = df.to_csv(sep='\t', index=False)
            st.download_button(
                label="📋 TSV 다운로드 (엑셀 붙여넣기용)",
                data=clipboard_text,
                file_name=f"test_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.tsv",
                mime="text/tab-separated-values",
                use_container_width=True
            )


if __name__ == "__main__":
    main()